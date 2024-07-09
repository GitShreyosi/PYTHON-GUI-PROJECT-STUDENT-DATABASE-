from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import *
from openpyxl.styles import Alignment,Font
from PIL import Image,ImageTk

def displayRecords():
    try:
        wb = load_workbook('database.xlsx')
    except FileNotFoundError:
        messagebox.showinfo("No Database", "No records")
        return

    def search_records():
        query = search_entry.get()
        search_entry.delete(0,END)

        found = False
        for row_id in treeview.get_children():
            if query.lower() == treeview.item(row_id)["values"][1].lower():
                treeview.selection_set(row_id)
                messagebox.showinfo("Record Found", "Record with Student's Name {} found.".format(treeview.item(row_id)["values"][1]))
                found = True
                break
        if not found:
            messagebox.showinfo("Record Not Found", "Record with Student's Name {} not found.".format(query))
            tree.selection_clear()
        return

    def delete_record():
        name = delete_entry.get()
        delete_entry.delete(0, END)

        file = 'database.xlsx'
        wb = load_workbook(file)
        sheet = wb.active
        found = False

        for row_id in treeview.get_children():
            if name.lower() == treeview.item(row_id)["values"][1].lower():
                opt = messagebox.askyesno("Confirmation","Are you sure?")
                if opt:
                    treeview.delete(row_id)
                    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=2, max_col=2):
                        for cell in row:
                            if cell.value is not None and cell.value.lower() == name.lower():
                                sheet.delete_rows(cell.row)
                                found = True
                                break
                        if found:
                            break

                    if found:
                        messagebox.showinfo("Record Deleted", "Record with Student's Name {} deleted.".format(cell.value))
                        wb.save(file)
                        return
                return

        messagebox.showinfo("Record Not Found", "Record with Student's Name {} not found.".format(name))
        wb.save(file)
        return

    def go_back():
        display_window.withdraw()
        home.deiconify()
        home.state('zoomed')
        return
       
    home.withdraw()
    display_window=Toplevel()
    display_window.state('zoomed')
    display_window.config(bg="#FFD39B")
    display_window.title("Display Records")
    display_window.geometry("2000x1000")
    canvas_4=Canvas(display_window)
    icon_image1=ImageTk.PhotoImage(Image.open("win.png"))
    canvas_4.create_image(0,0,anchor=NW,image=icon_image1)
    display_window.iconphoto(False, icon_image1)
    display_window.resizable(True, True)

    tree=ttk.Treeview(display_window)

    # Add search and filter functionality
    search_frame=Frame(display_window)
    search_frame.pack(pady=5)
    search_frame.config(bg="#FFD39B")
    searchlabel=Label(search_frame, text="Search by Student's Name:",font=("Calibri",14),bg="#FFD39B",fg="Blue")
    searchlabel.grid(row=0,column=0,padx=5,pady=5,sticky=W)
    search_entry=Entry(search_frame,width=50,font=("Calibri",12))
    search_entry.grid(row=0,column=1,padx=5,pady=5)

    search_button = Button(search_frame, text="SEARCH", command=search_records,width=20,font=("Calibri",12),bg="#00C957")
    search_button.grid(row=0,column=2,padx=5,pady=5)

    # Add delete functionality
    delete_label=Label(search_frame, text="Delete Record by Student's Name:",font=("Calibri",14),bg="#FFD39B",fg="Blue")
    delete_label.grid(row=1,column=0 ,padx=5,pady=5,sticky=W)
    delete_entry=Entry(search_frame,width=50,font=("Calibri",12))
    delete_entry.grid(row=1,column=1,padx=5,pady=5)

    delete_button = Button(search_frame, text="DELETE", command=delete_record,width=20,font=("Calibri",12),bg="#EE2C2C")
    delete_button.grid(row=1,column=2,padx=5)

    back_button=Button(display_window,text="BACK", command=go_back,width=15,font=("Calibri",12),bg="#6495ED")
    back_button.pack(pady=5)

    #Creating the treeview

    wb = load_workbook('database.xlsx')
    sheet = wb.active

    treeFrame = ttk.Frame(display_window)

    y_treeScroll = ttk.Scrollbar(treeFrame)
    y_treeScroll.pack(side="right",fill='y')

    x_treeScroll = ttk.Scrollbar(treeFrame,orient="horizontal")
    x_treeScroll.pack(side="bottom",fill='x')

    cols = (sheet['A3'],sheet['B3'],sheet['C3'],sheet['D3'],sheet['E3'],sheet['F3'],sheet['G3'],sheet['H3'],sheet['I3'],sheet['J3'])
    treeview = ttk.Treeview(treeFrame,show="headings",yscrollcommand=y_treeScroll.set,xscrollcommand=x_treeScroll.set,columns=cols,height=50)

    treeview.column(cols[0],width=190)
    treeview.column(cols[1],width=150)
    treeview.column(cols[2],width=320)
    treeview.column(cols[3],width=170)
    treeview.column(cols[4],width=110)
    treeview.column(cols[5],width=70)
    treeview.column(cols[6],width=160)
    treeview.column(cols[7],width=130)
    treeview.column(cols[8],width=150)
    treeview.column(cols[9],width=150)

    treeview.pack()
    y_treeScroll.config(command=treeview.yview)
    x_treeScroll.config(command=treeview.xview)

    style = ttk.Style()
    style.configure("Treeview", font=('Times New Roman', 12))  # Change the font size here

    # Reading the data from Excel file
    headers = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    # Create column indices for headings
    column_indices = list(range(len(headers)))

    for col_index, col_name in zip(column_indices, headers):
        treeview.heading(col_index, text=col_name, anchor=CENTER)
        treeview.column(col_index, anchor=CENTER)  # Center align column headers

    style.configure("Treeview.Heading", font=('Calibri', 14, 'bold'))  # Change the font style here

    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=10):
        data = [cell.value for cell in row]
        treeview.insert("", "end", values=data)

    treeFrame.pack(expand=True, fill="y",padx=5,pady=5)
    return

def recordEntry():
    home.withdraw()
    window1 = Toplevel()
    window1.title("Data Entry")
    window1.geometry("560x533+500+100")
    canvas_5=Canvas(window1)
    icon_image2=ImageTk.PhotoImage(Image.open("win.png"))
    canvas_5.create_image(0,0,anchor=NW,image=icon_image2)
    window1.iconphoto(False, icon_image)
    window1.resizable(False,False)
    root=Frame(window1,bg="#326273")
    root.pack()

    def submission():
        file = 'database.xlsx'
        try:
            wb = load_workbook(file)
        except FileNotFoundError:
            wb=Workbook()
            sheet=wb.active

            sheet.merge_cells('A1:J2')
            sheet['A1']="Students' details"
            sheet['A1'].alignment=Alignment(horizontal='center',vertical='center')
            sheet['A1'].font=Font(size=16,bold=True)
            sheet['A3']="Registration number"
            sheet['B3']="Student's name"
            sheet['C3']="Student's address"
            sheet['D3']="Guardian's name"
            sheet['E3']="Department"
            sheet['F3']="DOB"
            sheet['G3']="Email ID"
            sheet['H3']="Phone number"
            sheet['I3']="Class 10 marks(%)"
            sheet['J3']="Class 12 marks(%)"

            wb.save(file)
    
        sheet=wb.active

        # Find the maximum reg. no.
        registration_numbers = []
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    registration_numbers.append(cell.value)
        max_registration_number = max(registration_numbers, default=10000)
    
        sheet.cell(column=2,row=sheet.max_row+1,value=Entry1.get())
        sheet.cell(column=3,row=sheet.max_row,value=Entry2.get(1.0,END))
        sheet.cell(column=4,row=sheet.max_row,value=Entry3.get())
        sheet.cell(column=5,row=sheet.max_row,value=Entry4.get())
        sheet.cell(column=6,row=sheet.max_row,value=Entry5.get())
        sheet.cell(column=7,row=sheet.max_row,value=Entry6.get())
        sheet.cell(column=8,row=sheet.max_row,value=Entry7.get())
        sheet.cell(column=9,row=sheet.max_row,value=Entry8.get())
        sheet.cell(column=10,row=sheet.max_row,value=Entry9.get())

        # Generate new registration numbers greater than the maximum value
        for row_index in range(4, sheet.max_row + 1):
            current_cell = sheet.cell(row=row_index, column=1)
            if current_cell.value is None:
                max_registration_number+=1
                current_cell.value = max_registration_number

        wb.save(file)

        Entry1.delete(0,END)
        Entry2.delete(1.0,END)
        Entry3.delete(0,END)
        Entry4.delete(0,END)
        Entry5.delete(0,END)
        Entry6.delete(0,END)
        Entry7.delete(0,END)
        Entry8.delete(0,END)
        Entry9.delete(0,END)

        result = messagebox.askyesno("Submitted successfully","Do you want to add more records?")
        if result is False:
            window1.withdraw()
            home.deiconify()
        return
    
    def back():
        window1.withdraw()
        home.deiconify()
        home.state('zoomed')
        return

    head = Label(root,text="Student Registration",font=("Calibri",25),fg="#BDFCC9",bg="#326273").grid(row=0,column=0,columnspan=2,pady=5)

    Label1 = Label(root,text="Student's name:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=1,column=0,padx=5,pady=5,sticky=W)
    Entry1 = Entry(root,width=40,font=("Calibri",13))
    Entry1.grid(row=1,column=1,padx=5,pady=5)

    Label2 = Label(root,text="Student's address:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=2,column=0,padx=5,pady=5,sticky=W)
    Entry2 = Text(root,width=40,height=3,font=("Calibri",13))
    Entry2.grid(row=2,column=1,padx=5,pady=5)

    Label3 = Label(root,text="Guardian's name:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=3,column=0,padx=5,pady=5,sticky=W)
    Entry3 = Entry(root,width=40,font=("Calibri",13))
    Entry3.grid(row=3,column=1,padx=5,pady=5)

    Label4 = Label(root,text="Department:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=4,column=0,padx=5,pady=5,sticky=W)
    Entry4 = ttk.Combobox(root,width=38,values=['CSE','IT','ECE','EE','AIML'],font=("Calibri",13))
    Entry4.grid(row=4,column=1,padx=5,pady=5)

    Label5 = Label(root,text="DOB:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=5,column=0,padx=5,pady=5,sticky=W)
    Entry5 = DateEntry(root,width=38,font=("Calibri",13))
    Entry5.grid(row=5,column=1,padx=5,pady=5)

    Label6 = Label(root,text="Email ID:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=6,column=0,padx=5,pady=5,sticky=W)
    Entry6 = Entry(root,width=40,font=("Calibri",13))
    Entry6.grid(row=6,column=1,padx=5,pady=5)

    Label7 = Label(root,text="Phone number:",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=7,column=0,padx=5,pady=5,sticky=W)
    Entry7 = Entry(root,width=40,font=("Calibri",13))
    Entry7.grid(row=7,column=1,padx=5,pady=5)

    Label8 = Label(root,text="Class 10 marks(%):",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=8,column=0,padx=5,pady=5,sticky=W)
    Entry8 = Entry(root,width=40,font=("Calibri",13))
    Entry8.grid(row=8,column=1,padx=5,pady=5)

    Label9 = Label(root,text="Class 12 marks(%):",bg="#326273",fg="#FDF5E6",font=("Calibri",15)).grid(row=9,column=0,padx=5,pady=5,sticky=W)
    Entry9 = Entry(root,width=40,font=("Calibri",13))
    Entry9.grid(row=9,column=1,padx=5,pady=5)

    Button_submit = Button(root,text="SUBMIT",width=7,bg="LightGreen",command=submission,font=("Calibri",15))
    Button_submit.grid(row=10,column=0,pady=15,padx=5,sticky=W)

    Button_back = Button(root,text="BACK",width=7,fg="Red",bg="Yellow",command=back,font=("Calibri",15))
    Button_back.grid(row=10,column=1,pady=15,padx=5,sticky=E)
    return

def close():
    home.destroy()
    return

home = Tk()
home.state('zoomed')
home.geometry("2000x1000")
home.title("Student Information System")
# home.iconbitmap(default= 'download.png')

canvas_3=Canvas(home)
icon_image=ImageTk.PhotoImage(Image.open("win.png"))
canvas_3.create_image(0,0,anchor=NW,image=icon_image)
home.iconphoto(False, icon_image)




frame = Frame(master=home, width=2000,height=10000)
frame.pack()

canvas_3=Canvas(frame,width= 2000, height=10000)
bg_image_2=ImageTk.PhotoImage(Image.open("trial_1.jpg"))
canvas_3.create_image(0,0,anchor=NW,image=bg_image_2)

bg_label= Label(frame, image=bg_image_2)
bg_label.place(x=0, y=5,relheight=1,relwidth=1)


canvas_1=Canvas(frame,width=225,height=215)
bg_image=ImageTk.PhotoImage(Image.open("download.png"))
canvas_1.create_image(0,0,anchor=NW,image=bg_image)
canvas_1.grid(pady=30, padx= 30, column= 0)



canvas_2=Canvas(frame,width=225,height=212)
bg_image_1=ImageTk.PhotoImage(Image.open("Makaut.png"))
canvas_2.create_image(0,0,anchor=NW,image=bg_image_1)
canvas_2.grid(pady=30, padx= 30, column= 16, row =0)

 

# College label
Clg_label = Label(frame, text = "St. Thomas' College of Engineering and Technology", font = ('Algerian', 25, "bold"),background="Black", fg="Yellow")
Clg_label.grid(row=0, column=1, columnspan=10)

Clg_label = Label(frame, text = "STUDENT INFORMATION", font = ('Times New Roman', 22, "bold"),background="Black", fg="Yellow")
Clg_label.grid(row=2, column=1, columnspan=10)

add = Button(frame, text = 'ADD', padx = 30, pady= 20,background= 'Yellow', fg='Red',font= ('Helvetica', 15, "bold"), command=recordEntry)
add.grid(row=10,column=0,pady=390)

display = Button(frame, text = 'DISPLAY', padx = 30, pady= 20,background= 'Cyan', fg='Black',font= ('Helvetica', 15, "bold"), command=displayRecords)
display.grid(row=10, column=5,pady=390)

exit = Button(frame, text = 'EXIT', padx = 30, pady= 20,background= 'Yellow', fg='Red',font= ('Helvetica', 15,"bold"), command=close)
exit.grid(row = 10, column =16,pady=390)


home.mainloop()
